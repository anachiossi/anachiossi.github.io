"""
build_geo_json.py
Reads _data/ana_chiossi_data_clean.xlsx, geocodes missing locations via Nominatim,
writes lat/lon back into the Excel file, and writes _output/geo.json.
Only run when adding new filming locations.
Double-click to run on Windows.
"""

import pandas as pd
import json
import os
import sys
import time
from openpyxl import load_workbook

os.chdir(os.path.dirname(os.path.abspath(__file__)))

DATA_FILE   = "_data/ana_chiossi_data_clean.xlsx"
OUTPUT_FILE = "_output/geo.json"
NOMINATIM_UA = "ana_chiossi_portfolio_site"

def read(sheet):
    df = pd.read_excel(DATA_FILE, sheet_name=sheet, header=1)
    return df.iloc[0:].reset_index(drop=True)

def safe_float(val):
    if pd.isna(val):
        return None
    s = str(val).strip()
    if not s or s in ("nan", " "):
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None

def geocode(location_city, city, country):
    try:
        from urllib.request import urlopen, Request
        from urllib.parse import urlencode
        import json as _json

        loc = str(location_city).strip() if location_city and str(location_city).strip() not in ("", "nan") else ""
        query = f"{loc}, {city}, {country}" if loc else f"{city}, {country}"

        params = urlencode({"q": query, "format": "json", "limit": 1})
        url = f"https://nominatim.openstreetmap.org/search?{params}"
        req = Request(url, headers={"User-Agent": NOMINATIM_UA})

        with urlopen(req, timeout=10) as resp:
            results = _json.loads(resp.read().decode())

        if results:
            return float(results[0]["lat"]), float(results[0]["lon"])

        # fallback to city level
        if loc:
            params2 = urlencode({"q": f"{city}, {country}", "format": "json", "limit": 1})
            url2 = f"https://nominatim.openstreetmap.org/search?{params2}"
            req2 = Request(url2, headers={"User-Agent": NOMINATIM_UA})
            with urlopen(req2, timeout=10) as resp2:
                results2 = _json.loads(resp2.read().decode())
            if results2:
                print(f"       (fell back to city level)")
                return float(results2[0]["lat"]), float(results2[0]["lon"])

        return None, None

    except Exception as e:
        print(f"       geocoding error: {e}")
        return None, None

print("\n" + "="*60)
print("  BUILD GEO JSON")
print("="*60)

if not os.path.exists(DATA_FILE):
    print(f"\n  File not found: {DATA_FILE}")
    input("\nPress Enter to close...")
    sys.exit(1)

os.makedirs("_output", exist_ok=True)

print("\n  Reading sheets...")

geo_finder = read("geo_localization_finder")
film_geo   = read("film_geo")
fd         = read("film_details")

# existing coords from film_geo
existing_coords = {}
for _, row in film_geo.iterrows():
    _id = str(row.get("_id", "")).strip()
    lat = safe_float(row.get("lat"))
    lon = safe_float(row.get("lon"))
    if _id and _id != "nan" and lat is not None and lon is not None:
        existing_coords[_id] = (lat, lon)

# manual overrides + needs geocoding from geo_localization_finder
manual_coords   = {}
needs_geocoding = []

for _, row in geo_finder.iterrows():
    _id = str(row.get("_id", "")).strip()
    if not _id or _id == "nan":
        continue
    lat = safe_float(row.get("lat"))
    lon = safe_float(row.get("lon"))
    city    = str(row.get("city", "")).strip()
    country = str(row.get("country", "")).strip()
    loc     = str(row.get("location_city", "")).strip()

    if lat is not None and lon is not None:
        manual_coords[_id] = (lat, lon)
    elif city and country:
        needs_geocoding.append((_id, loc, city, country))

# geocode missing
new_coords = {}
if needs_geocoding:
    print(f"\n  Geocoding {len(needs_geocoding)} missing location(s)...")
    print("  (1 second pause per Nominatim rules)\n")
    for _id, loc, city, country in needs_geocoding:
        display = f"{city}, {country}" + (f" ({loc})" if loc and loc != "nan" else "")
        print(f"  -> {_id}  {display}")
        lat, lon = geocode(loc, city, country)
        if lat is not None and lon is not None:
            new_coords[_id] = (lat, lon)
            print(f"       lat: {round(lat,6)}, lon: {round(lon,6)}")
        else:
            print(f"       NOT FOUND - add coordinates manually")
        time.sleep(1)
else:
    print("\n  No new locations to geocode.")

# write new coords back to Excel
if new_coords:
    print(f"\n  Writing {len(new_coords)} coordinate(s) back to Excel...")
    wb = load_workbook(DATA_FILE)

    for sheet_name in ["film_geo", "geo_localization_finder"]:
        ws = wb[sheet_name]
        id_col = lat_col = lon_col = None
        for cell in ws[2]:
            v = str(cell.value).strip() if cell.value else ""
            if v == "_id":
                id_col = cell.column
            elif v == "lat":
                lat_col = cell.column
            elif v == "lon":
                lon_col = cell.column

        if id_col and lat_col and lon_col:
            for row in ws.iter_rows(min_row=3):
                cid = row[id_col-1].value
                if cid and str(cid).strip() in new_coords:
                    _id = str(cid).strip()
                    row[lat_col-1].value = new_coords[_id][0]
                    row[lon_col-1].value = new_coords[_id][1]

    wb.save(DATA_FILE)
    print(f"  Excel updated with new coordinates")

# build geo.json - priority: manual > new > existing
all_coords = {**existing_coords, **new_coords, **manual_coords}

fd_map = {}
for _, row in fd.iterrows():
    _id = str(row.get("_id", "")).strip()
    if _id and _id != "nan":
        r1 = str(row.get("role_1", "")).strip()
        fd_map[_id] = {
            "title":        str(row.get("title", "")).strip(),
            "type":         str(row.get("type", "")).strip(),
            "department":   str(row.get("department", "")).strip(),
            "role":         r1 if r1 and r1 != "nan" else None,
            "release_year": row.get("release_year"),
        }

locations = []
no_coords = []

for _id, (lat, lon) in sorted(all_coords.items()):
    film = fd_map.get(_id, {})
    ry = film.get("release_year")
    try:
        year = int(float(str(ry))) if pd.notna(ry) else None
    except:
        year = None

    locations.append({
        "id":           _id,
        "title":        film.get("title", _id),
        "type":         film.get("type"),
        "department":   film.get("department"),
        "role":         film.get("role"),
        "release_year": year,
        "lat":          round(lat, 6),
        "lon":          round(lon, 6),
    })

for _id in fd_map:
    if _id not in all_coords:
        no_coords.append(_id)

with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
    json.dump(locations, f, ensure_ascii=False, indent=2)

print(f"\n  {len(locations)} locations written to {OUTPUT_FILE}")
if new_coords:
    print(f"  {len(new_coords)} new location(s) geocoded")
if no_coords:
    print(f"\n  {len(no_coords)} film(s) with no coordinates:")
    for nid in sorted(no_coords):
        print(f"       {nid}  {fd_map.get(nid,{}).get('title', nid)}")

print(f"\n  Check _output/geo.json before copying to assets/")
print("="*60)
input("\nPress Enter to close...")
